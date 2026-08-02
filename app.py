import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning, module="langchain_community")

import os
import shutil
import subprocess
import re
import json
import queue
import sys
from pathlib import Path
from dotenv import load_dotenv
from langchain_ollama import ChatOllama
from langchain.agents import create_agent
from langchain.tools import tool
from langchain_community.tools import DuckDuckGoSearchRun

from docx import Document
from openpyxl import Workbook
from fpdf import FPDF

# Librerías para voz
import pyaudio
import vosk
import pyttsx3
import time
import speech_recognition as sr  # Se mantiene para casos con internet

load_dotenv()

# ----------------------------------------------------------------------
# RUTAS PERSONALIZADAS DE TUS APLICACIONES
# ----------------------------------------------------------------------
RUTAS_PERSONALIZADAS = {
    "whatsapp": "shell:AppsFolder\\WhatsAppDesktop",
    "linkedin": r"C:\Program Files\WindowsApps\7EE7776C.LinkedInforWindows_3.0.43.0_x64__w1wdnht996qgy\LinkedIn.exe",
    "vscode": r"C:\Users\HOME\AppData\Local\Programs\Microsoft VS Code\bin\code.cmd",
    "chrome": r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    "edge": r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    "bloc de notas": r"C:\Windows\System32\notepad.exe",
    "calculadora": r"C:\Windows\System32\calc.exe",
    "explorador": r"C:\Windows\explorer.exe",
    "cmd": r"C:\Windows\System32\cmd.exe",
    "powershell": r"C:\Windows\System32\WindowsPowerShell\v1.0\powershell.exe",
}

# ----------------------------------------------------------------------
# CONFIGURACIÓN DEL MODELO
# ----------------------------------------------------------------------
llm = ChatOllama(model="llama3.2", temperature=0.1)

# ----------------------------------------------------------------------
# FUNCIÓN PARA RESOLVER RUTAS
# ----------------------------------------------------------------------
def resolver_ruta(ruta: str) -> str:
    ruta_original = ruta.strip()
    patron_unidad = re.compile(
        r'(?:nuevo vol\s*\(([a-z])\)|disco\s*([a-z])|([a-z]):)',
        re.IGNORECASE
    )
    match = patron_unidad.search(ruta_original)
    if match:
        letra = match.group(1) or match.group(2) or match.group(3)
        if letra:
            letra = letra.upper()
            resto = ruta_original[match.end():].lstrip("\\/ ")
            return f"{letra}:\\{resto}"
    especiales = {
        "escritorio": os.path.join(os.environ["USERPROFILE"], "Desktop"),
        "documentos": os.path.join(os.environ["USERPROFILE"], "Documents"),
        "descargas": os.path.join(os.environ["USERPROFILE"], "Downloads"),
        "música": os.path.join(os.environ["USERPROFILE"], "Music"),
        "imágenes": os.path.join(os.environ["USERPROFILE"], "Pictures"),
        "videos": os.path.join(os.environ["USERPROFILE"], "Videos"),
        "inicio": os.environ["USERPROFILE"],
    }
    ruta_limpia = ruta_original.lower().strip()
    for clave, valor in especiales.items():
        if ruta_limpia.startswith(clave):
            resto = ruta_original[len(clave):].lstrip("\\/ ")
            return os.path.join(valor, resto)
    return ruta_original

# ----------------------------------------------------------------------
# FUNCIÓN PARA BUSCAR EJECUTABLES
# ----------------------------------------------------------------------
def buscar_ejecutable(nombre: str) -> str | None:
    nombre_limpio = re.sub(r'^(abrir|el|la|los|las)\s+', '', nombre, flags=re.IGNORECASE).strip().lower()
    for clave, ruta in RUTAS_PERSONALIZADAS.items():
        if nombre_limpio == clave.lower():
            if ruta.startswith("shell:"):
                return ruta
            elif os.path.isfile(ruta):
                return ruta
            else:
                return None
    try:
        resultado = subprocess.run(
            f'where {nombre_limpio}',
            capture_output=True,
            text=True,
            shell=True
        )
        if resultado.returncode == 0 and resultado.stdout:
            rutas = resultado.stdout.strip().split('\n')
            if rutas:
                return rutas[0].strip()
    except:
        pass
    carpetas_busqueda = [
        "C:\\Program Files",
        "C:\\Program Files (x86)",
        os.environ["LOCALAPPDATA"] + "\\Programs",
        os.environ["USERPROFILE"] + "\\AppData\\Local\\Programs",
        os.environ["USERPROFILE"] + "\\AppData\\Local\\Microsoft\\WindowsApps",
    ]
    extensiones = [".exe", ".cmd", ".bat", ""]
    for carpeta in carpetas_busqueda:
        if not os.path.exists(carpeta):
            continue
        for ext in extensiones:
            ruta_candidata = os.path.join(carpeta, nombre_limpio + ext)
            if os.path.isfile(ruta_candidata):
                return ruta_candidata
        try:
            for item in os.listdir(carpeta):
                subcarpeta = os.path.join(carpeta, item)
                if os.path.isdir(subcarpeta):
                    for ext in extensiones:
                        ruta_candidata = os.path.join(subcarpeta, nombre_limpio + ext)
                        if os.path.isfile(ruta_candidata):
                            return ruta_candidata
        except:
            continue
    return None

# ----------------------------------------------------------------------
# FUNCIONES AUXILIARES DE EJECUCIÓN
# ----------------------------------------------------------------------
def ejecutar_borrado(ruta: str) -> str:
    try:
        ruta_obj = Path(ruta)
        if ruta_obj.is_file():
            ruta_obj.unlink()
        elif ruta_obj.is_dir():
            shutil.rmtree(ruta_obj)
        else:
            return f"No se encontró: {ruta}"
        return f"✅ Borrado exitoso: {ruta}"
    except Exception as e:
        return f"❌ Error al borrar: {str(e)}"

def ejecutar_movido(origen: str, destino: str) -> str:
    try:
        shutil.move(origen, destino)
        return f"✅ Movido exitosamente: {origen} → {destino}"
    except Exception as e:
        return f"❌ Error al mover: {str(e)}"

def _vaciar_papelera() -> str:
    try:
        subprocess.run(
            ["powershell", "-Command", "Clear-RecycleBin -Force"],
            capture_output=True,
            check=True,
            text=True
        )
        return "✅ Papelera de reciclaje vaciada."
    except subprocess.CalledProcessError as e:
        return f"❌ Error al vaciar la papelera: {e.stderr.strip() if e.stderr else str(e)}"
    except Exception as e:
        return f"❌ Error inesperado: {str(e)}"

# ----------------------------------------------------------------------
# HERRAMIENTAS DEL AGENTE
# ----------------------------------------------------------------------

@tool(description="Crea una nueva carpeta en la ruta especificada. Entiende 'escritorio', 'documentos', 'D:', etc.")
def crear_carpeta(ruta: str) -> str:
    ruta_real = resolver_ruta(ruta)
    try:
        Path(ruta_real).mkdir(parents=True, exist_ok=True)
        return f"✅ Carpeta creada en: {ruta_real}"
    except Exception as e:
        return f"❌ Error: {str(e)}"

@tool(description="Abre una aplicación instalada (WhatsApp, Chrome, Word, etc.) buscando en rutas personalizadas y en el sistema.")
def abrir_programa(nombre: str) -> str:
    nombre_original = nombre.strip()
    for clave, valor in RUTAS_PERSONALIZADAS.items():
        if nombre_original.lower() == clave.lower():
            ruta = valor
            if ruta.startswith("shell:"):
                try:
                    subprocess.Popen(f"start {ruta}", shell=True)
                    return f"✅ {nombre_original} abierto (identificador de la Store)."
                except Exception as e:
                    return f"❌ Error al abrir: {str(e)}"
            elif os.path.isfile(ruta):
                try:
                    subprocess.Popen(ruta, shell=True)
                    return f"✅ {nombre_original} abierto desde: {ruta}"
                except Exception as e:
                    return f"❌ Error al abrir: {str(e)}"
            else:
                return f"❌ La ruta para '{nombre_original}' no existe: {ruta}"
    ruta_encontrada = buscar_ejecutable(nombre_original)
    if ruta_encontrada and not ruta_encontrada.startswith("shell:"):
        try:
            subprocess.Popen(ruta_encontrada, shell=True)
            return f"✅ {nombre_original} abierto desde: {ruta_encontrada}"
        except Exception as e:
            return f"❌ Error al abrir: {str(e)}"
    elif ruta_encontrada and ruta_encontrada.startswith("shell:"):
        try:
            subprocess.Popen(f"start {ruta_encontrada}", shell=True)
            return f"✅ {nombre_original} abierto (identificador)."
        except Exception as e:
            return f"❌ Error al abrir: {str(e)}"
    else:
        return f"❌ No se encontró '{nombre_original}'."

@tool(description="Busca información en internet usando DuckDuckGo. Útil para preguntas de actualidad, noticias, o cualquier cosa que requiera información externa.")
def buscar_internet(consulta: str) -> str:
    """Realiza una búsqueda en DuckDuckGo y devuelve los resultados resumidos."""
    try:
        search = DuckDuckGoSearchRun()
        resultado = search.invoke(consulta)
        if len(resultado) > 1000:
            resultado = resultado[:1000] + "..."
        return resultado
    except Exception as e:
        return f"❌ Error en la búsqueda (puede que no tengas internet): {str(e)}"

@tool(description="Abre un archivo con su aplicación asociada.")
def abrir_archivo(ruta: str) -> str:
    ruta_real = resolver_ruta(ruta)
    if not os.path.exists(ruta_real):
        return f"❌ El archivo no existe: {ruta_real}"
    try:
        os.startfile(ruta_real)
        return f"✅ Archivo abierto: {ruta_real}"
    except Exception as e:
        return f"❌ Error: {str(e)}"

@tool(description="Lista los archivos y carpetas en la ruta especificada.")
def listar_archivos(ruta: str = ".") -> str:
    ruta_real = resolver_ruta(ruta)
    try:
        elementos = os.listdir(ruta_real)
        if not elementos:
            return "La carpeta está vacía."
        return "\n".join(elementos)
    except Exception as e:
        return f"❌ Error: {str(e)}"

@tool(description="Solicita confirmación para borrar un archivo o carpeta.")
def borrar_archivo(ruta: str) -> str:
    ruta_real = resolver_ruta(ruta)
    return f"__CONFIRMAR_BORRADO__:{ruta_real}"

@tool(description="Solicita confirmación para mover un archivo o carpeta.")
def mover_archivo(origen: str, destino: str) -> str:
    origen_real = resolver_ruta(origen)
    destino_real = resolver_ruta(destino)
    return f"__CONFIRMAR_MOVIMIENTO__:{origen_real}|{destino_real}"

@tool(description="Crea un documento (txt, docx, xlsx, pdf) en la ruta especificada.")
def crear_documento(tipo: str, ruta: str, contenido: str = "") -> str:
    ruta_real = resolver_ruta(ruta)
    try:
        Path(ruta_real).parent.mkdir(parents=True, exist_ok=True)
        if tipo.lower() == "txt":
            with open(ruta_real, "w", encoding="utf-8") as f:
                f.write(contenido if contenido else "Documento de texto")
            return f"✅ TXT creado: {ruta_real}"
        elif tipo.lower() == "docx":
            doc = Document()
            doc.add_paragraph(contenido if contenido else "Documento Word")
            doc.save(ruta_real)
            return f"✅ DOCX creado: {ruta_real}"
        elif tipo.lower() == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws["A1"] = contenido if contenido else "Documento Excel"
            wb.save(ruta_real)
            return f"✅ XLSX creado: {ruta_real}"
        elif tipo.lower() == "pdf":
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            if contenido:
                pdf.multi_cell(0, 10, contenido)
            else:
                pdf.cell(200, 10, txt="Documento PDF", ln=True)
            pdf.output(ruta_real)
            return f"✅ PDF creado: {ruta_real}"
        else:
            return f"❌ Tipo '{tipo}' no soportado."
    except Exception as e:
        return f"❌ Error: {str(e)}"

@tool(description="Vacía completamente la papelera de reciclaje de Windows.")
def vaciar_papelera() -> str:
    return _vaciar_papelera()

# ----------------------------------------------------------------------
# CREAR EL AGENTE
# ----------------------------------------------------------------------
tools = [
    crear_carpeta,
    abrir_programa,
    buscar_internet,
    abrir_archivo,
    listar_archivos,
    borrar_archivo,
    mover_archivo,
    crear_documento,
    vaciar_papelera
]
agent = create_agent(llm, tools)

# ----------------------------------------------------------------------
# GESTIÓN DE CONFIRMACIONES
# ----------------------------------------------------------------------
def procesar_mensaje(mensaje: str) -> str:
    response = agent.invoke({"messages": [("human", mensaje)]})
    contenido = response['messages'][-1].content

    if "__CONFIRMAR_BORRADO__" in contenido:
        ruta = contenido.split(":", 1)[1]
        print(f"\n⚠️ El agente quiere BORRAR: {ruta}")
        confirmar = input("¿Autorizas el borrado? (sí/no): ").strip().lower()
        if confirmar in ["sí", "si", "s", "yes", "y"]:
            resultado = ejecutar_borrado(ruta)
            return f"Confirmación recibida. {resultado}"
        else:
            return "Borrado cancelado por el usuario."

    elif "__CONFIRMAR_MOVIMIENTO__" in contenido:
        _, datos = contenido.split(":", 1)
        origen, destino = datos.split("|", 1)
        print(f"\n⚠️ El agente quiere MOVER:")
        print(f"   Origen: {origen}")
        print(f"   Destino: {destino}")
        confirmar = input("¿Autorizas el movimiento? (sí/no): ").strip().lower()
        if confirmar in ["sí", "si", "s", "yes", "y"]:
            resultado = ejecutar_movido(origen, destino)
            return f"Confirmación recibida. {resultado}"
        else:
            return "Movimiento cancelado por el usuario."

    else:
        return contenido

# ----------------------------------------------------------------------
# FUNCIONES DE VOZ (CON VOSK PARA OFFLINE)
# ----------------------------------------------------------------------
# Variable global para el modelo Vosk (se carga una sola vez)
vosk_model = None

def cargar_modelo_vosk():
    global vosk_model
    # Ruta donde descargaste el modelo de español
    model_path = "C:\\Users\\HOME\\vosk-model-small-es-0.42"  # Cambia si lo pusiste en otro lado
    if not os.path.exists(model_path):
        print("⚠️ No se encontró el modelo Vosk. Descárgalo de: https://alphacephei.com/vosk/models")
        print("   Colócalo en:", model_path)
        return False
    try:
        vosk_model = vosk.Model(model_path)
        print("✅ Modelo Vosk cargado correctamente.")
        return True
    except Exception as e:
        print(f"❌ Error al cargar modelo Vosk: {e}")
        return False

def escuchar() -> str:
    """Escucha usando Vosk (offline) si está disponible, o falla a speech_recognition."""
    global vosk_model
    
    # Si no hay modelo Vosk, intentar cargarlo
    if vosk_model is None:
        if not cargar_modelo_vosk():
            # Si falla, usar speech_recognition con internet
            print("⚠️ Usando speech_recognition (requiere internet).")
            return escuchar_con_google()
    
    # Usar Vosk
    recognizer = vosk.KaldiRecognizer(vosk_model, 16000)
    pa = pyaudio.PyAudio()
    
    try:
        # Abrir el micrófono con los parámetros correctos
        stream = pa.open(
            format=pyaudio.paInt16,
            channels=1,
            rate=16000,
            input=True,
            frames_per_buffer=4000
        )
        print("\n🎤 Escuchando... (di algo)")
        stream.start_stream()
        
        # Buffer para acumular audio
        audio_data = b""
        silence_frames = 0
        max_silence = 10  # frames de silencio para considerar fin de frase
        
        while True:
            data = stream.read(4000, exception_on_overflow=False)
            audio_data += data
            
            # Verificar si el reconocedor tiene resultado
            if recognizer.AcceptWaveform(data):
                result = json.loads(recognizer.Result())
                texto = result.get("text", "")
                if texto:
                    stream.stop_stream()
                    stream.close()
                    pa.terminate()
                    print(f"Tú: {texto}")
                    return texto
            else:
                # Obtener resultado parcial (para detectar silencio)
                partial = json.loads(recognizer.PartialResult())
                partial_text = partial.get("partial", "")
                if partial_text:
                    # Si hay silencio durante un tiempo, considerar fin de frase
                    if not partial_text:
                        silence_frames += 1
                    else:
                        silence_frames = 0
                    
                    # Si lleva mucho tiempo en silencio, finalizar
                    if silence_frames > max_silence:
                        stream.stop_stream()
                        stream.close()
                        pa.terminate()
                        print("⏱️ Tiempo de espera agotado.")
                        return ""
        return ""
    except Exception as e:
        print(f"❌ Error en Vosk: {e}")
        return ""
    finally:
        try:
            stream.stop_stream()
            stream.close()
            pa.terminate()
        except:
            pass

def escuchar_con_google() -> str:
    """Fallback: usar speech_recognition con Google (requiere internet)."""
    recognizer = sr.Recognizer()
    with sr.Microphone() as source:
        print("\n🎤 Escuchando... (di algo)")
        recognizer.adjust_for_ambient_noise(source, duration=1)
        try:
            audio = recognizer.listen(source, timeout=5, phrase_time_limit=10)
            print("🔄 Procesando...")
            texto = recognizer.recognize_google(audio, language='es-ES')
            print(f"Tú: {texto}")
            return texto
        except sr.WaitTimeoutError:
            print("⏱️ Tiempo de espera agotado.")
            return ""
        except sr.UnknownValueError:
            print("🤔 No entendí lo que dijiste.")
            return ""
        except sr.RequestError as e:
            print(f"❌ Error de conexión (necesitas internet para este modo): {e}")
            return ""

# Motor de voz (se reinicia en cada llamada)
def hablar(texto: str):
    print(f"🔊 Agente (voz): {texto}")
    try:
        engine = pyttsx3.init()
        engine.setProperty('rate', 150)
        engine.setProperty('volume', 0.9)
        engine.say(texto)
        engine.runAndWait()
        time.sleep(0.2)
        engine.stop()
    except Exception as e:
        print(f"❌ Error al hablar: {e}")

# ----------------------------------------------------------------------
# BUCLE PRINCIPAL
# ----------------------------------------------------------------------
if __name__ == "__main__":
    print("🤖 Agente con superpoderes ")
    print("Escribe 'salir' para terminar.\n")
    modo = input("¿Usar VOZ (v) o TECLADO (t)? ").strip().lower()
    usar_voz = modo == 'v'

    if usar_voz:
        # Intentar cargar Vosk al inicio
        cargar_modelo_vosk()
        hablar("Hola , soy tu asistente. ¿En qué puedo ayudarte?")
    else:
        print("Modo teclado activado.")

    while True:
        if usar_voz:
            pregunta = escuchar()
            if not pregunta:
                continue
            if pregunta.lower() in ["gracias hablamos después", "cállate", "quit", "adiós", "bye"]:
                hablar("¡Hasta luego!")
                break
        else:
            pregunta = input("Tú: ")
            if pregunta.lower() in ["gracias hablamos después", "cállate", "quit", "adiós", "bye"]:
                print("👋 ¡Hasta luego!")
                break

        # --- DETECCIÓN DIRECTA DE "abre la URL" ---
        if pregunta.lower().startswith("abre la url") or pregunta.lower().startswith("abre la página"):
            partes = pregunta.split(maxsplit=3)
            if len(partes) >= 4:
                url = partes[3]
            else:
                url = "https://www.google.com"
            try:
                subprocess.Popen(f"start {url}", shell=True)
                respuesta = f"✅ URL abierta: {url}"
            except Exception as e:
                respuesta = f"❌ Error al abrir URL: {str(e)}"
            print(f"Agente: {respuesta}")
            if usar_voz:
                hablar(respuesta)
            continue

        # --- DETECCIÓN DIRECTA DE PAPELERA ---
        if "papelera" in pregunta.lower() or "reciclaje" in pregunta.lower():
            respuesta = _vaciar_papelera()
            print(f"Agente: {respuesta}")
            if usar_voz:
                hablar(respuesta)
            continue

        # --- RESTO DE PREGUNTAS ---
        respuesta = procesar_mensaje(pregunta)
        print(f"Agente: {respuesta}")
        if usar_voz:
            hablar(respuesta)